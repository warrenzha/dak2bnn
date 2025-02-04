import os
import sys
import argparse
import time
import pickle
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

import numpy as np
import matplotlib.pyplot as plt
from matplotlib.markers import MarkerStyle
from matplotlib.ticker import MaxNLocator, AutoMinorLocator
from sklearn.model_selection import KFold
import pandas as pd

import torch

from dak.utils.util import MinMaxNormalize, PrintOutput
from dak.utils.metrics import RegressionMetrics
from dataloader import UCIDatasets

from nn_uci import NNUCI
from nnsvgp_uci import NNSVGPUCI
from svdkl_uci import SVDKLUCI
from avdkl_uci import AVDKLUCI
from dak_uci import DAKUCI

# # Turn on cudNN benchmarking
# torch.backends.cudnn.benchmark = True

# Get directory paths
dir_path = os.path.dirname(os.path.abspath(__file__))  # /dak/examples/uci
checkpoint_dir = os.path.join(dir_path, "checkpoint")  # /dak/examples/uci/checkpoint
checkpoint_dataloader_dir = os.path.join(checkpoint_dir, f'dataloader')# /dak/examples/uci/checkpoint/dataloader
data_dir = os.path.join(dir_path, "data")  # /dak/examples/uci/data
data_basemetrics_dir = os.path.join(data_dir, f'basemetrics') # /dak/examples/uci/data/basemetrics
data_totalmetrics_dir = os.path.join(data_dir, f'totalmetrics') # /dak/examples/uci/data/totalmetrics
figs_dir = os.path.join(dir_path, "figs")  # /dak/examples/uci/figs


# Create directories if they don't exist
os.makedirs(checkpoint_dir, exist_ok=True)
os.makedirs(checkpoint_dataloader_dir, exist_ok=True)
os.makedirs(data_dir, exist_ok=True)
os.makedirs(data_basemetrics_dir, exist_ok=True)
os.makedirs(data_totalmetrics_dir, exist_ok=True)
os.makedirs(figs_dir, exist_ok=True)

# enable cuda
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")


# Create argument parser
parser = argparse.ArgumentParser(description='UCI')

# Add arguments
parser.add_argument('--running-datasets', 
                    type=str,
                    nargs='+',
                    default=['wine', 'gas', 'parkinsons', 'kin40k', 'protein', 'kegg'],
                    choices=['wine', 'gas', 'parkinsons', 'kin40k', 'protein', 'kegg'], 
                    help='Choose the datasets to use.')
parser.add_argument('--model-names', 
                    type=str,
                    nargs='+',
                    default=['nn', 'nnsvgp', 'svdkl', 'avdkl', 'dak-mc', 'dak-cf'], 
                    choices=['nn', 'nnsvgp', 'svdkl', 'avdkl', 'dak-mc', 'dak-cf'], 
                    help='Choose the DKL models to use.')
parser.add_argument('--metrics', 
                    type=str,
                    nargs='+',
                    default=['train_times', 'test_times', 'rmses', 'nlpds', 'coverage_scores'], 
                    choices=['train_times', 'test_times', 'rmses', 'nlpds', 'coverage_scores'], 
                    help='Choose the metrics to create plots and tables.')
parser.add_argument('--dnn-name', 
                    type=str,
                    default='fnn', 
                    choices=['fnn', 'resnet1d'], 
                    help='Choose the dnn models to use.')
parser.add_argument('-j',
                    '--workers',
                    default=2,
                    type=int,
                    metavar='N',
                    help='number of data loading workers (default: 2)')
parser.add_argument('--epochs',
                    default=100,
                    type=int,
                    metavar='N',
                    help='number of total epochs to run')
parser.add_argument('-b',
                    '--batch-sizes',
                    nargs='+',
                    default=[512],#[512, 1024],
                    type=int,
                    metavar='N',
                    help='mini-batch size (default: 256)')
parser.add_argument('--nn-out-features',
                    nargs='+',
                    default=[16, 64, 256],
                    type=int,
                    metavar='N',
                    help='number of output features of NN for NN model (default: 16)')
parser.add_argument('--hidden-features',
                    default=[64, 32],
                    nargs='+',
                    type=int,
                    metavar='N',
                    help='a list of the number of two hidden featrues in NN (default: [128, 64])')
parser.add_argument('--num-proj',
                    default=8,
                    type=int,
                    metavar='N',
                    help='number of projections (base GP layers) (default: 8)')
parser.add_argument('--lr',
                    '--learning-rate',
                    default=0.001,
                    type=float,
                    metavar='LR',
                    help='initial learning rate')
parser.add_argument('--weight-decay',
                    '--wd',
                    default=5e-4,
                    type=float,
                    metavar='W',
                    help='weight decay (default: 5e-4)')
parser.add_argument('--half',
                    dest='half',
                    action='store_true',
                    help='use half-precision(16-bit) ')
parser.add_argument('--num-mc-train',
                    type=int,
                    default=8,
                    metavar='N',
                    help='number of Monte Carlo runs during training')
parser.add_argument('--num-mc-test',
                    type=int,
                    default=20,
                    metavar='N',
                    help='number of Monte Carlo samples to be drawn during inference')
parser.add_argument('--seed',
                    type=int,
                    default=10,
                    metavar='S',
                    help='random seed (default: 10)')
parser.add_argument('--kfolds',
                    default=5,
                    type=int,
                    metavar='N',
                    help='number of test sets for evaluation (default: 5)')
parser.add_argument('--num-ip',
                    default=64,
                    type=int,
                    metavar='N',
                    help='number of inducing points for SVGP layer (default: 64)')
parser.add_argument('--grid-size',
                    default=256,
                    type=int,
                    metavar='N',
                    help='number of inducing points (grid_size) for KISS-GP in SVDKL (default: 128)')
parser.add_argument('--split-ratio',
                    default=0.8,
                    type=float,
                    metavar='SR',
                    help='split ratio of train/test (default: 0.8)')
parser.add_argument('--test-split-ratio',
                    default=0.6,
                    type=float,
                    metavar='SR',
                    help='ratio of subtest/test (default: 0.6)')
parser.add_argument('--noise-var',
                    default=0.01,
                    type=float,
                    metavar='NV',
                    help='noise variance (default: 0.01)')
parser.add_argument('--device',
                    default=device,
                    metavar='DV',
                    help='device')
parser.add_argument('-vb',
                    '--verbose', 
                    default=True,
                    type=bool,
                    action=argparse.BooleanOptionalAction,
                    help='if print out logs; (default: True)')
parser.add_argument('-val',
                    '--validate',
                    default=False,
                    type=bool,
                    action=argparse.BooleanOptionalAction,
                    help='if validate model on validation set; (default: False)')
parser.add_argument('--test-on-train',
                    default=True,
                    type=bool,
                    action=argparse.BooleanOptionalAction,
                    help='if test on training set; (default: True)')
parser.add_argument('--transform',
                    default=False,
                    type=bool,
                    action=argparse.BooleanOptionalAction,
                    help='if transform the data to [0,1]; (default: False)')

args = parser.parse_args()


transform = MinMaxNormalize() if args.transform else None

def dataloader_setup(dataset, batch_size):
    # K-Fold generate
    kfold = KFold(n_splits=args.kfolds, shuffle=True, random_state=args.seed)
    
    # train_size = int(args.split_ratio * len(dataset))
    # test_size = len(dataset) - train_size

    # # Split the dataset
    # train_dataset, test_dataset = torch.utils.data.random_split(
    #     dataset, [train_size, test_size], 
    #     generator=torch.Generator().manual_seed(args.seed),
    # )

    # Wrap the dataset with a transform
    # train_dataset.dataset.transform = transform
    # test_dataset.dataset.transform = transform

    train_loaders = [0]*args.kfolds # a list [0,...,0] of size kfolds
    test_loaders = [0]*args.kfolds # a list [0,...,0] of size kfolds
    for fold_id, (train_ids, test_ids) in enumerate(kfold.split(dataset)):
        # Sample elements randomly from a given list of ids, no replacement.
        train_subsampler = torch.utils.data.SubsetRandomSampler(train_ids)
        test_subsampler = torch.utils.data.SubsetRandomSampler(test_ids)
        
        # Define data loaders for training and testing data in this fold
        train_loaders[fold_id] = torch.utils.data.DataLoader(
                        dataset, batch_size=batch_size, 
                        sampler=train_subsampler, 
                        num_workers=args.workers,
                        pin_memory=True,
                    )
        test_loaders[fold_id] = torch.utils.data.DataLoader(
                        dataset, batch_size=batch_size, 
                        sampler=test_subsampler, 
                        num_workers=args.workers,
                        pin_memory=True,
                    )

    return train_loaders, test_loaders



def experiment_setup(dataset, batch_size, nn_out_features, model_name):

    ###################################
    # Initialize zero torch.Tensors to store the metrics
    ###################################
    kfolds = args.kfolds
    rmses, nlpds, coverage_scores, \
    train_times, test_times, \
    train_rmses, train_nlpds, train_coverage_scores = [
        torch.zeros(kfolds, device=args.device) for _ in range(8)
    ]
    train_losses = torch.zeros(kfolds, args.epochs, device=args.device)
    test_outs_list = [0]*kfolds
    train_outs_list = [0]*kfolds

    ###################################
    # Model intialize
    ###################################
    if model_name == 'nn':
        UCIModel = NNUCI(args, nn_out_features=nn_out_features)
    elif model_name == 'nnsvgp':
        UCIModel = NNSVGPUCI(args, nn_out_features=nn_out_features)
    elif model_name == 'svdkl':
        UCIModel = SVDKLUCI(args, nn_out_features=nn_out_features)
    elif model_name == 'avdkl':
        UCIModel = AVDKLUCI(args, nn_out_features=nn_out_features)
    elif model_name == 'dak-mc':
        UCIModel = DAKUCI(args, mc_sampling=True, nn_out_features=nn_out_features, batch_size=batch_size)
    elif model_name == 'dak-cf':
        UCIModel = DAKUCI(args, mc_sampling=False, nn_out_features=nn_out_features, batch_size=batch_size)

    ###################################
    # Model setup
    ###################################
    UCIModel.model_setup()
    model = UCIModel.model  # Initialize the model before either training or loading

    ###################################
    # Setup train and test dataloaders
    ###################################
    dataloaderpath = os.path.join(checkpoint_dataloader_dir, 
                                f'b{batch_size}_{dataset.__name__}_seed{args.seed}')

    if os.path.exists(dataloaderpath):
        # Load
        with open(dataloaderpath , 'rb') as file:
            train_loaders, test_loaders = pickle.load(file)
    else:
        train_loaders, test_loaders = dataloader_setup(dataset, batch_size)
        save_tuple = train_loaders, test_loaders
        # Save
        with open(dataloaderpath, 'wb') as file:
            pickle.dump(save_tuple, file)
        print(f"Dataloaders for seed = {args.seed} saved successfully.")

    ###################################
    # Loop over kfolds
    ###################################
    for kfold_id, (train_loader, test_loader) in enumerate(zip(train_loaders, test_loaders)):
        
        ######################
        # Checkpoint path
        ######################
        checkpoint_kfold_dir = os.path.join(checkpoint_dir, f'{kfold_id}_kfold')
        os.makedirs(checkpoint_kfold_dir, exist_ok=True)
        filename = f'b{batch_size}_p{nn_out_features}_{model_name}_{args.dnn_name}_{dataset.__name__}.pth'
        checkpoint_path = os.path.join(checkpoint_kfold_dir, filename)

        

        ###################################
        # Train and save (Or load the trained model)
        ###################################
        if not os.path.exists(checkpoint_path):
            # Train and save model
            start = time.time()
            model, train_losses_list = UCIModel.train(train_loader, val_loader=test_loaders[0])
            train_times[kfold_id] = time.time() - start
            train_losses[kfold_id, :] = torch.as_tensor(train_losses_list, device=args.device)
            # Save model state dict and attributes (train_times, train_losses)
            save_checkpoint = {
                'model_state_dict': model.state_dict(),
                'train_times': train_times,
                'train_losses': train_losses,
                'args': args  # save args too
            }
            torch.save(save_checkpoint, checkpoint_path)
        else:
            # Load model
            checkpoint = torch.load(checkpoint_path, weights_only=False)
            model.load_state_dict(checkpoint['model_state_dict'])

            # Load the attributes
            train_times = checkpoint.get('train_times', None)  # Get train_times if available
            train_losses = checkpoint.get('train_losses', None)  # Get train_losses if available
            # args = checkpoint.get('args', None)  # Optionally load saved args

        ###################################
        # Test on test_loader
        ###################################
        # # loop over randomly selected test subsets
        # for kfold_id, test_loader in enumerate(test_loaders):
        start = time.time()
        # pred_mean, pred_std, input_true, target_true = UCIModel.test(test_loader, model)
        test_outs = UCIModel.test(test_loader, model)
        test_times[kfold_id] = time.time() - start
        test_outs_list[kfold_id] = test_outs

        # measure metrics
        reg_metrics = RegressionMetrics(pred_mean=test_outs[0], 
                                        pred_var=test_outs[1]**2, 
                                        test_y=test_outs[3])
        rmses[kfold_id] = reg_metrics.rmse()
        nlpds[kfold_id] = reg_metrics.nlpd()
        coverage_scores[kfold_id] = reg_metrics.coverage_score(num_std=1)

        if args.test_on_train:
            ###################################
            # Test on train_loader
            ###################################
            train_outs = UCIModel.test(train_loader, model)
            train_outs_list[kfold_id] = train_outs

            # measure metrics
            train_metrics = RegressionMetrics(pred_mean=train_outs[0], 
                                            pred_var=train_outs[1]**2, 
                                            test_y=train_outs[3])
            train_rmses[kfold_id] = train_metrics.rmse()
            train_nlpds[kfold_id] = train_metrics.nlpd()
            train_coverage_scores[kfold_id] = train_metrics.coverage_score(num_std=1)

        if args.verbose:
            print(f"\033[95m {dataset.__name__.upper()} \033[00m")
            print(f"\033[91m ==============={model_name.upper()}=============== \033[00m")
            print(f"\033[91m batch_size = {batch_size}, nn_out_features = {nn_out_features} \033[00m")
            print(f"\033[92m {kfold_id}-Fold \033[00m")
            
            print(f"Train Time = {train_times[kfold_id]}")
            print(f"Test Time = {test_times[kfold_id]}")
            
            print(f"\033[92m --------------Test set-------------- \033[00m")
            print(f"RMSE = {rmses[kfold_id]}")
            print(f"NLPD = {nlpds[kfold_id]}")
            print(f"Coverage Score = {coverage_scores[kfold_id]}")

            print(f"\033[92m --------------Train set-------------- \033[00m")
            print(f"RMSE = {train_rmses[kfold_id]}")
            print(f"NLPD = {train_nlpds[kfold_id]}")
            print(f"Coverage Score = {train_coverage_scores[kfold_id]}")



    ###################################
    # Save the useful data to a dict for (batch_size, nn_out_features, model_name)
    ###################################

    save_dict = {}
    keys = [
        'train_times', 'test_times',
        'rmses','nlpds', 'coverage_scores', 'train_losses',
        'dataset_name', 'batch_size', 'nn_out_features', 'model_name',
        'train_rmses', 'train_nlpds', 'train_coverage_scores',
        'train_outs', 'test_outs'
        ]
    vals = [
        train_times, test_times,
        rmses, nlpds, coverage_scores, train_losses,
        str(dataset.__name__), batch_size, nn_out_features, model_name,
        train_rmses, train_nlpds, train_coverage_scores,
        train_outs_list, test_outs_list
        ]

    for (key, val) in zip(keys, vals):
        save_dict[key] = val

    return save_dict


def main():
    # running_datasets = ['wine', 'gas', 'parkinsons', 'kin40k', 'protein', 'kegg']
    for dataset_name in args.running_datasets:
        # filepath to load and save
        filename = f"metrics_{dataset_name}.pkl"
        filepath = os.path.join(data_totalmetrics_dir, filename) # Specify the filename
        
        if os.path.exists(filepath):
            # Load
            with open(filepath , 'rb') as file:
                save_dict = pickle.load(file)
            print(f"Metrics for {dataset_name} data loaded successfully.")

        else:
            dataset = UCIDatasets(dataset=dataset_name, transform=transform)
            save_dict = {}
            for batch_size in args.batch_sizes:
                if batch_size not in save_dict:
                    save_dict[batch_size] = {}
                for nn_out_features in args.nn_out_features:
                    if nn_out_features not in save_dict[batch_size]:
                        save_dict[batch_size][nn_out_features] = {}
                    for model_name in args.model_names:
                        
                        # save or load the base_dict
                        base_dict_path = os.path.join(
                            data_basemetrics_dir, 
                            f"b{batch_size}_p{nn_out_features}_{model_name}_{dataset_name}.pkl"
                        )
                        if os.path.exists(base_dict_path):
                            with open(base_dict_path , 'rb') as file:
                                base_dict = pickle.load(file)
                        else:
                            base_dict = experiment_setup(
                                dataset, batch_size, nn_out_features, model_name
                            )
                            with open(base_dict_path, 'wb') as file:
                                pickle.dump(base_dict, file)

                        # Print out Metrics
                        if args.verbose:
                            print(f"\033[95m {dataset.__name__.upper()} \033[00m")
                            print(f"\033[91m ==============={model_name.upper()}=============== \033[00m")
                            print(f"\033[91m batch_size = {batch_size}, nn_out_features = {nn_out_features} \033[00m")
                            # print(f"\033[92m Loop over {args.kfolds}-Folds \033[00m")
                            
                            print(f"Train Time = {base_dict['train_times'].mean()} ± {base_dict['train_times'].std()}")
                            print(f"Test Time = {base_dict['test_times'].mean()} ± {base_dict['test_times'].std()}")
                            
                            print(f"\033[92m --------------Test set-------------- \033[00m")
                            print(f"RMSE = {base_dict['rmses'].mean()} ± {base_dict['rmses'].std()}")
                            print(f"NLPD = {base_dict['nlpds'].mean()} ± {base_dict['nlpds'].std()}")
                            print(f"Coverage Score = {base_dict['coverage_scores'].mean()} ± {base_dict['coverage_scores'].std()}")

                            print(f"\033[92m --------------Train set-------------- \033[00m")
                            print(f"RMSE = {base_dict['train_rmses'].mean()} ± {base_dict['train_rmses'].std()}")
                            print(f"NLPD = {base_dict['train_nlpds'].mean()} ± {base_dict['train_nlpds'].std()}")
                            print(f"Coverage Score = {base_dict['train_coverage_scores'].mean()} ± {base_dict['train_coverage_scores'].std()}")

                        save_dict[batch_size][nn_out_features][model_name] = base_dict

            # Save
            with open(filepath, 'wb') as file:
                pickle.dump(save_dict, file)
            print(f"Metrics for {dataset_name} data saved successfully.")


def save_to_excel(
        dataset_names,
        batch_sizes=args.batch_sizes,
        nn_out_features=args.nn_out_features,
        model_names=args.model_names, 
        metrics=['train_times', 'test_times', 'rmses', 'nlpds', 'coverage_scores']):
    # Define a mapping from internal metric names to human-readable names
    metric_name_mapping = {
        'train_times': 'Train Time',
        'test_times': 'Test Time',
        'rmses': 'RMSE',
        'nlpds': 'NLPD',
        'coverage_scores': 'Coverage Rate',
    }

     # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Step 1: Generate the multi-level header
    # Row 1 (Dataset, N, D, Model + Batch Size)
    ws.merge_cells("A1:A3")
    ws.merge_cells("B1:B3")
    ws.merge_cells("C1:C3")
    ws.merge_cells("D1:D3")
    ws["A1"] = "Dataset"
    ws["B1"] = "N"
    ws["C1"] = "D"
    ws["D1"] = "Model"
    
    current_col = 5
    for batch_size in batch_sizes:
        nn_out_feature_count = len(nn_out_features) * len(metrics)
        ws.merge_cells(start_row=1, start_column=current_col, 
                       end_row=1, end_column=current_col + nn_out_feature_count - 1)
        ws.cell(row=1, column=current_col).value = f"BS={batch_size}"
        
        for nn_out in nn_out_features:
            ws.merge_cells(start_row=2, start_column=current_col, 
                           end_row=2, end_column=current_col + len(metrics) - 1)
            ws.cell(row=2, column=current_col).value = f"NN={nn_out}"
            
            for metric in metrics:
                ws.cell(row=3, column=current_col).value = metric_name_mapping.get(metric, metric)
                current_col += 1

    # Step 2: Populate the data for each dataset
    current_row = 4  # Start after the header

    for dataset_name in dataset_names:
        dataset = UCIDatasets(dataset_name)
        N = dataset.x.shape[0]  # Get N (number of samples)
        D = dataset.x.shape[1]  # Get D (number of features)
        
        # Apply the required formatting for dataset_name
        formatted_dataset_name = dataset_name.upper() if dataset_name == 'kegg' else dataset_name.capitalize()

        # Load the metrics dictionary for the dataset
        filename = f"metrics_{dataset_name}.pkl"
        filepath = os.path.join(data_totalmetrics_dir, filename)
        with open(filepath, 'rb') as file:
            save_dict = pickle.load(file)

        # Track the starting row for merging Dataset, N, and D
        dataset_start_row = current_row
        model_start_row = current_row

        # For each model_name
        for model_name in model_names:
            row_data = [formatted_dataset_name, N, D, model_name.upper()]

            # Append the metrics data for each batch_size and nn_out_feature
            for batch_size in batch_sizes:
                for nn_out in nn_out_features:
                    for metric in metrics:
                        val = save_dict[batch_size][nn_out][model_name][metric]
                        mean_ = val.mean().detach().cpu().numpy()
                        std_ = val.std().detach().cpu().numpy()
                        row_data.append(f"{mean_:.3f} ± {std_:.3f}")

            # Write the row data to the worksheet
            ws.append(row_data)
            current_row += 1

        # Merge cells for Dataset, N, and D columns across the same dataset
        ws.merge_cells(start_row=dataset_start_row, start_column=1, end_row=current_row - 1, end_column=1)  # Dataset column
        ws.merge_cells(start_row=dataset_start_row, start_column=2, end_row=current_row - 1, end_column=2)  # N column
        ws.merge_cells(start_row=dataset_start_row, start_column=3, end_row=current_row - 1, end_column=3)  # D column

    # Apply center alignment to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Adjust column widths
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Save to an Excel file
    output_filename = "uci_metrics_excel.xlsx"
    output_filepath = os.path.join(data_dir, output_filename)
    wb.save(output_filepath)
    print(f"Excel table saved to {output_filepath}")


def save_to_latex(
        dataset_names,
        batch_sizes=args.batch_sizes,
        nn_out_features=args.nn_out_features,
        model_names=args.model_names,
        metrics=['train_times', 'test_times', 'rmses', 'nlpds', 'coverage_scores']):
    # Define a mapping from internal metric names to human-readable names
    metric_name_mapping = {
        'train_times': 'Train Time',
        'test_times': 'Test Time',
        'rmses': 'RMSE',
        'nlpds': 'NLPD',
        'coverage_scores': 'Coverage Rate',
    }

    latex_code = "\\begin{table*}[ht]\n\\centering\n\\caption{UCI Metrics Table}\n\\resizebox{\\linewidth}{!}{%\n"
    latex_code += "\\begin{tabular}{|l|l" + "|".join(["c"] * (len(batch_sizes) * len(nn_out_features) * len(metrics))) + "|}\n"
    latex_code += "\\hline\n"

    # Step 1: Multi-level header (Dataset + Batch Size, NN out features, Metrics)
    latex_code += "\\multirow{3}{*}{\\textbf{Dataset (N,D)}} & \\multirow{3}{*}{\\textbf{Model}}"
    
    # Row 1 (Batch Size)
    for batch_size in batch_sizes:
        latex_code += f" & \\multicolumn{{{len(nn_out_features) * len(metrics)}}}{{c|}}{{\\textbf{{BS={batch_size}}}}}"
    latex_code += " \\\\\n\\cline{3-" + str(2 + len(batch_sizes) * len(nn_out_features) * len(metrics)) + "}\n"

    # Row 2 (NN Out Features)
    latex_code += "& "
    for batch_size in batch_sizes:
        for nn_out in nn_out_features:
            latex_code += f" & \\multicolumn{{{len(metrics)}}}{{c|}}{{\\textbf{{NN={nn_out}}}}}"
    latex_code += " \\\\\n\\cline{3-" + str(2 + len(batch_sizes) * len(nn_out_features) * len(metrics)) + "}\n"

    # Row 3 (Metrics)
    latex_code += "& "
    for batch_size in batch_sizes:
        for nn_out in nn_out_features:
            for metric in metrics:
                latex_code += f" & \\textbf{{{metric_name_mapping.get(metric, metric)}}}"
    latex_code += " \\\\\n\\hline\n"

    # Step 2: Populate the data for each dataset
    for dataset_name in dataset_names:
        dataset = UCIDatasets(dataset_name)
        N = dataset.x.shape[0]  # Get N (number of samples)
        D = dataset.x.shape[1]  # Get D (number of features)

        # Apply the required formatting for dataset_name, and add (N,D)
        formatted_dataset_name = dataset_name.upper() if dataset_name == 'kegg' else dataset_name.capitalize()
        dataset_info = f"{formatted_dataset_name} \\\\ (N={N}, D={D})"

        # Load the metrics dictionary for the dataset
        filename = f"metrics_{dataset_name}.pkl"
        filepath = os.path.join(data_totalmetrics_dir, filename)
        with open(filepath, 'rb') as file:
            save_dict = pickle.load(file)

        # For each model_name
        first_row = True
        for model_name in model_names:
            if first_row:
                latex_code += f"\\multirow{{{len(model_names)}}}{{*}}{{{dataset_info}}} & {model_name.upper()}"
                first_row = False
            else:
                latex_code += f"& {model_name.upper()}"

            # Append the metrics data for each batch_size and nn_out_feature
            for batch_size in batch_sizes:
                for nn_out in nn_out_features:
                    for metric in metrics:
                        val = save_dict[batch_size][nn_out][model_name][metric]
                        mean_ = val.mean().detach().cpu().numpy()
                        std_ = val.std().detach().cpu().numpy()
                        latex_code += f" & {mean_:.3f} $\\pm$ {std_:.3f}"

            latex_code += " \\\\\n"

        # Add a horizontal line between datasets
        latex_code += "\\hline\n"

    latex_code += "\\end{tabular}%\n}\n\\end{table*}"

    # Save the LaTeX code to a .tex file
    output_filename = "uci_metrics_latex.tex"
    output_filepath = os.path.join(data_dir, output_filename)
    with open(output_filepath, 'w') as latex_file:
        latex_file.write(latex_code)

    print(f"LaTeX table code saved to {output_filepath}")


def barplot(dataset_names=['wine', 'gas']):
    
    metrics = ['train_time', 'test_times', 'rmses', 'nlpds']
    titles = ['Training time (seconds)', 'Inference time (seconds)', 'RMSE', 'NLPD']

    width = 0.15 # width of each bar

    model_names = ['nn', 'nnsvgp', 'svdkl', 'avdkl', 'dak-mc', 'dak-cf']
    labels = ['NN', 'NN+SVGP', 'SV-DKL', 'AV-DKL', r'$\bf{DAK-MC}$', r'$\bf{DAK-CF}$']
    xlabels = [dataset_name.upper() if dataset_name == 'kegg' else dataset_name.capitalize() for dataset_name in dataset_names]

    palette_colors = plt.get_cmap('tab10').colors
    colors = palette_colors[:len(model_names)]


    x = np.arange(len(dataset_names)) # X-axis positions for each group of bars

    for (metric, title) in zip(metrics, titles):
        # Create the plot
        fig, ax = plt.subplots(figsize=(12, 8))

        for i, (model_name, color, label) in enumerate(zip(model_names, colors, labels)):
            temp_metrics = [0]*len(dataset_names)
            yerr = [0]*len(dataset_names)
            for dataset_id, dataset_name in enumerate(dataset_names):
                filename = f"metrics_{dataset_name}_dict.pkl"
                filepath = os.path.join(data_dir, filename)
                with open(filepath , 'rb') as file:
                    save_dict = pickle.load(file)
                
                val = save_dict[model_name][metric]
                temp_metrics[dataset_id] = val if metric=='train_time' else val.mean().detach().cpu().numpy()
                yerr[dataset_id] = np.nan if metric=='train_time' else val.std().detach().cpu().numpy()
            
            # Define error bar properties
            error_kw = {
                'elinewidth': 2.5,  # Thicker error bars
                'ecolor': 'black',  # Error bar color
                'capsize': 10,      # Cap size at the end of error bars
                'capthick': 2       # Cap thickness
            }


            # Plot bars for each method
            x_position = x + i * width - (width * (len(model_names) / 2))
            ax.bar(x_position, temp_metrics, width, label=label, color=color,
                   yerr=yerr, align='center', alpha=0.8, error_kw=error_kw)

        ax.set_xlabel('Dataset', fontsize=28)
        # ax.set_ylabel('Training time (s)', fontsize=16)
        ax.set_title(title, fontsize=32, fontweight='bold')


        ax.tick_params(axis='x', labelsize=20)
        ax.tick_params(axis='y', labelsize=20)

        ax.set_xticks(x)
        ax.set_xticklabels(xlabels, rotation=0, ha='right', fontsize=28)

        # Use MaxNLocator to fix the number of Y-ticks (5 major ticks)
        ax.yaxis.set_major_locator(MaxNLocator(nbins=8, integer=False))
        ax.yaxis.set_minor_locator(AutoMinorLocator())
        ax.grid(True, which='major', linestyle=':', linewidth=0.7, color='gray')
        ax.grid(True, which='minor', linestyle=':', linewidth=0.4, color='lightgray')
        # Move the legend below the entire plot
        fig.legend(
            loc='upper center', bbox_to_anchor=(0.5+0.05, 0.05), 
            ncol=len(model_names), fontsize=20,
        )

        plt.tight_layout(rect=[0.05, 0.05, 1, 0.95])
        plt.close()

        fig_path = os.path.join(figs_dir, f"{metric}_barplot.png")
        fig.savefig(fig_path, bbox_inches='tight')

def loss_plot(dataset_names=['wine', 'gas', 'parkinsons']):
    metric = 'train_losses'
    model_names = ['nn', 'nnsvgp', 'svdkl', 'avdkl', 'dak-mc', 'dak-cf'][-2:]
    labels = ['NN', 'NN+SVGP', 'SV-DKL', 'AV-DKL', 'DAK-MC', 'DAK-CF'][-2:]

    sub_titles = [dataset_name.upper() if dataset_name == 'kegg' else dataset_name.capitalize() for dataset_name in dataset_names]


    palette_colors = plt.get_cmap('tab10').colors
    markers_class = list(MarkerStyle.markers.keys())

    colors = palette_colors[:len(model_names)]
    markers = markers_class[0 : 0+len(model_names)*2:2]

    ls = '-'
    lw = 3.5
    alpha = 1.0
    markersize = 10

    nrows, ncols = 2, 3
    fig, axs = plt.subplots(nrows, ncols, figsize=(12*ncols, 8*nrows))

    grid_x = np.arange(1, args.epochs+1)
    tick_positions = np.arange(len(grid_x)+1)

    i = 0
    j = 0
    for (dataset_name, sub_title) in zip(dataset_names, sub_titles):
        ax = axs[i,j]

        filename = f"metrics_{dataset_name}_dict.pkl"
        filepath = os.path.join(data_dir, filename)
        with open(filepath , 'rb') as file:
            save_dict = pickle.load(file)

        for (model_name, color, marker, label) in zip(model_names, colors, markers, labels):
            metric_val= np.array(save_dict[model_name][metric])
            # metric_val = np.log1p(metric_val)
            ax.plot(grid_x, metric_val,
                    color=color, ls=ls, lw=lw, alpha=alpha,
                    marker=marker, markersize=markersize, label=label)

        ax.set_xlabel("Epoch", fontsize=40, labelpad=18)
        # ax.set_ylabel(f'{title_name}', fontsize = 40, fontweight='bold')
        ax.set_title(f'{sub_title}', pad=30, fontsize=45, fontweight='bold', color='blue')


        # Set x-ticks to be evenly spaced with values from grid_x
        ax.set_xticks(tick_positions)
        ax.set_xticklabels(tick_positions)
        ax.tick_params(axis='x', labelsize=33, rotation=0, length=10, width=2, colors='black', direction='inout')

        # Set tick parameters (size of xtick, ytick)
        ax.tick_params(axis='x', labelsize=25)
        ax.tick_params(axis='y', labelsize=25)

        if i == 0 and j == 0:
            ax.legend(fontsize=38)


        ax.xaxis.set_major_locator(MaxNLocator(nbins=11, integer=False))
        ax.yaxis.set_major_locator(MaxNLocator(nbins=8, integer=False))
        ax.yaxis.set_minor_locator(AutoMinorLocator())
        ax.grid(True, which='major', linestyle=':', linewidth=0.7, color='gray')
        ax.grid(True, which='minor', linestyle=':', linewidth=0.4, color='lightgray')

        if j < ncols-1:
            j += 1
        else:
            i += 1
            j = 0

    plt.subplots_adjust(bottom=0.1)
    plt.tight_layout(rect=[0.05, 0.05, 1, 0.95])
    fig.suptitle(f"Training loss", fontsize=65, fontweight='bold', y=1.02)
    plt.close()

    fig_path = os.path.join(figs_dir, f"{metric}_plot.png")
    fig.savefig(fig_path, bbox_inches='tight')


def fitplot(dataset_names=['wine', 'gas']):
    pass


if __name__ == "__main__":
    out_filename = os.path.join(data_dir, 'out.txt')
    with open(out_filename, 'w') as f:
        sys.stdout = PrintOutput(f) # Redirect stdout to both the terminal and the file
        main()
        sys.stdout = sys.__stdout__ # Restore stdout back to its original state if needed
        f.close()
    
    # save_to_excel(args.running_datasets, model_names=args.model_names, metrics=args.metrics)
    # save_to_latex(args.running_datasets, model_names=args.model_names, metrics=args.metrics)
    # barplot(dataset_names=args.running_datasets)
    # loss_plot(dataset_names=args.running_datasets)